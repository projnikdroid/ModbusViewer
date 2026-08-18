"""Convert a plant tag-list Excel workbook into ModbusViewer's sampletags.csv format.

The workbook's actual column names vary per export and are never hardcoded here —
a JSON mapping file declares, per sheet, which source column feeds which
sampletags.csv field. Source column names in the mapping must exactly match the
sheet's real header text (see tag_mapping.example.json). registerType and the
0-based PDU address are derived automatically from a standard 5-digit Modicon
address in the source data (0xxxx=Coil, 1xxxx=DiscreteInput, 3xxxx=InputRegister,
4xxxx=HoldingRegister) rather than being part of the column mapping.

Note: Excel numeric-typed cells drop leading zeros, so a Coil address like 00001
must be entered as text in the source sheet to survive as a 5-digit value.

Examples:
    python tools/excel_tags_to_csv.py --data plant_tags.xlsx --mapping tag_mapping.example.json --output sampletags.csv
"""

import argparse
import csv
import json

import openpyxl

OUTPUT_COLUMNS = [
    "label",
    "description",
    "registerType",
    "address",
    "format",
    "byteOrder",
    "scale",
    "offset",
    "unit",
]

TEXT_FIELDS = ("label", "description", "unit")

REGISTER_TYPE_BY_PREFIX = {
    "0": "Coil",
    "1": "DiscreteInput",
    "3": "InputRegister",
    "4": "HoldingRegister",
}


class RowError(Exception):
    """A single source row can't be converted; the caller skips it and warns."""


class MappingError(Exception):
    """The mapping file doesn't match the workbook; aborts the whole run."""


def _address_digits(raw) -> str:
    if isinstance(raw, bool):
        raise RowError(f"address {raw!r} is not numeric")
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float):
        if not raw.is_integer():
            raise RowError(f"address {raw!r} is not a whole number")
        return str(int(raw))
    text = str(raw).strip()
    if not text:
        raise RowError("address is empty")
    return text


def parse_modicon_address(raw) -> tuple[str, int]:
    digits = _address_digits(raw)
    if len(digits) != 5 or not digits.isdigit():
        raise RowError(f"address {raw!r} is not a 5-digit Modicon address")

    prefix, register_number = digits[0], int(digits[1:])
    register_type = REGISTER_TYPE_BY_PREFIX.get(prefix)
    if register_type is None:
        raise RowError(f"address {raw!r} has unrecognized register-type prefix {prefix!r}")
    if register_number < 1:
        raise RowError(f"address {raw!r} has register number below 1")

    return register_type, register_number - 1


def translate_format(raw_data_type, data_type_map: dict) -> str:
    key = str(raw_data_type).strip().lower()
    normalized_map = {name.strip().lower(): value for name, value in data_type_map.items()}
    if key not in normalized_map:
        raise RowError(f"unrecognized data type {raw_data_type!r}")

    value = normalized_map[key]
    return "" if value.upper() == "BOOL" else value


def sanitize_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace(",", ";").strip()


def validate_mapping_against_headers(sheet_headers: list, column_map: dict, sheet_name: str) -> None:
    missing = [source for source in column_map if source not in sheet_headers]
    if missing:
        raise MappingError(f"sheet {sheet_name!r} is missing mapped column(s): {', '.join(missing)}")


def build_row(sheet_row: dict, column_map: dict, data_type_map: dict, defaults: dict) -> dict:
    mapped = {output_field: sheet_row.get(source_name) for source_name, output_field in column_map.items()}

    if mapped.get("address") in (None, ""):
        raise RowError("row has no address value")
    register_type, address = parse_modicon_address(mapped["address"])

    row = dict(defaults)
    row["registerType"] = register_type
    row["address"] = address

    if mapped.get("format") not in (None, ""):
        row["format"] = translate_format(mapped["format"], data_type_map)

    for field in TEXT_FIELDS:
        if field in mapped:
            row[field] = sanitize_text(mapped[field])

    if not row.get("label"):
        raise RowError("row has no label value")

    for field in OUTPUT_COLUMNS:
        row.setdefault(field, "")

    return row


def _read_sheet_rows(sheet) -> tuple[list, list]:
    rows_iter = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if all(value is None for value in values):
            continue
        rows.append(dict(zip(headers, values)))
    return headers, rows


def convert_workbook(workbook, mapping: dict) -> tuple[list, list]:
    output_rows = []
    warnings = []
    data_type_map = mapping.get("data_type_map", {})
    defaults = mapping.get("defaults", {})

    for sheet_name, sheet_config in mapping["sheets"].items():
        if sheet_name not in workbook.sheetnames:
            raise MappingError(f"mapping references sheet {sheet_name!r} not present in workbook")

        sheet = workbook[sheet_name]
        column_map = sheet_config["columns"]
        headers, rows = _read_sheet_rows(sheet)
        validate_mapping_against_headers(headers, column_map, sheet_name)

        for row_index, sheet_row in enumerate(rows, start=2):
            try:
                output_rows.append(build_row(sheet_row, column_map, data_type_map, defaults))
            except RowError as exc:
                warnings.append(f"{sheet_name} row {row_index}: {exc}")

    return output_rows, warnings


def write_csv(rows: list, output_path: str) -> None:
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--data", required=True, help="path to the source .xlsx workbook")
    parser.add_argument("--mapping", required=True, help="path to the JSON mapping file")
    parser.add_argument("--output", required=True, help="path to write the generated sampletags.csv")
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    with open(args.mapping, encoding="utf-8") as f:
        mapping = json.load(f)

    workbook = openpyxl.load_workbook(args.data, data_only=True)
    rows, warnings = convert_workbook(workbook, mapping)

    for warning in warnings:
        print(f"warning: {warning}")

    write_csv(rows, args.output)
    print(f"wrote {len(rows)} tag(s) to {args.output} ({len(warnings)} row(s) skipped)")


if __name__ == "__main__":
    main()
